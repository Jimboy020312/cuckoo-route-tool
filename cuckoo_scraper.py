"""
Cuckoo+ Service Specialist — monthly list scraper (Appium / UiAutomator2)

CONFIRMED SCREEN FLOW (from real XML captures)
--------------------------------------------------
1. LIST screen: each customer is a card with label/value pairs (NS No,
   Sales No, NS Date, Status, Appt Date, Product Name, Cust Name) plus
   one button.
2. Tapping that button does NOT open the detail screen directly — it
   opens a "Choose Option" POPUP with several choices (View Order, CCS
   Note, Appointment, Contact List, Cancel Appointment, Cancel). The
   script taps "View Order" from that popup.
3. That opens a "Customer Information" screen with TWO TABS:
     - "Address/Contact Info" (shown by default) — has THREE sections:
       Billing Address & Contact, Installation/Service Address &
       Contact, and Emergency Contact. The first two have normal
       label/value pairs PLUS a few fields with no visible label at
       all (customer name, a masked ID number, and the address) —
       these are identified by position instead of a label. Emergency
       Contact is fully labeled, no unlabeled fields.
     - "Sales Info" — reached by tapping its tab header. Normal
       label/value pairs, no unlabeled fields.
4. One driver.back() from the detail screen returns to the list.

HOW LABEL/VALUE PAIRING WORKS
---------------------------------
On every screen here, a label sits to the LEFT of its value, and both
share the same vertical position (y-coordinate) — but the raw reading
order in the XML doesn't alternate label-value cleanly, and the exact
pixel positions differ between devices/screen resolutions. So instead
of matching on fixed x-coordinates (which broke the first time this
was tested on a different device — a 720x1520 screen instead of the
1600x2560 one the fields were originally mapped from), the script
groups elements into rows by shared y-coordinate, then within each row
takes the leftmost text as the label and the rightmost as its value.
This holds regardless of screen size. Unlabeled fields (no matching
text in the known label list) are collected separately, in top-to-
bottom order, and assigned fixed names based on where they fall in
that section.

ONE THING WORTH NOTING
--------------------------
The list screen's "Product Name" field (e.g. "CP-XN501HW") is the
product model/SKU. The Sales Info tab's "Product" field (e.g. "XCEL")
is a different, separate product-name value from that section of the
app — both are genuine product names from different parts of the
record, not a mislabeled dealer code.

OUTPUT NOTE
--------------------------
Everything above and everything in the scraping/scrolling code below
is UNCHANGED — only the final export step is different. The exported
file only includes the 6 fields actually needed (sales_no, appt_date,
sales_info_product, install_contact_person, install_mobile1,
install_address), in that order, plus proposed_date (typed by hand),
a WhatsApp Message column (auto-fills once a date is typed, formatted
to match the WhatsApp bold style: *Hanis*, *Tarikh:*, *Alamat:*,
*Produk:*, *Nombor Pesanan:*), and a WhatsApp Link column — a tap-to-
send wa.me link with that message already filled in via URL, so
there's nothing to copy/paste by hand. This has to be .xlsx rather
than .csv, since a plain CSV can't hold a live formula.

Start with LIST_LIMIT = 3. Once the export looks right, set it to
None to process the entire list.
"""

import json
import os
import re
import time
import urllib.parse
import urllib.request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# ============================================================
# CONFIG
# ============================================================

APP_PACKAGE = "cuckoo.doctress"
# confirmed: the list screen
APP_ACTIVITY = "cuckoo.doctress.naturalcareservicelist"

LIST_LIMIT = None   # was 3 for testing — now processes the whole list
# set True again only if something breaks and you need to see raw element data
DEBUG = False

# --- List screen (confirmed from XML) ---
LABEL_FIELD_MAP = {
    "NS No": "ns_no",
    "Sales No": "sales_no",
    "NS Date": "ns_date",
    "Status": "status",
    "Appt Date": "appt_date",
    "Product Name": "product_name",
    "Cust Name": "cust_name",
}
KEY_FIELD = "ns_no"   # unique per card — avoids re-scraping someone after scrolling
# a card missing any of these is
REQUIRED_LIST_FIELDS = set(LABEL_FIELD_MAP.values())
# probably cut off by the screen
# edge, not actually incomplete data

# --- Detail screen: Address/Contact Info tab (confirmed from XML) ---
BILLING_LABEL_MAP = {
    "Doc No.": "billing_doc_no",
    "Sales No.": "billing_sales_no",
    "Contact Person": "billing_contact_person",
    "Tel No (Mobile 1)": "billing_mobile1",
    "Tel No (Mobile 2)": "billing_mobile2",
    "Tel No (Office)": "billing_office_phone",
    "Email": "billing_email",
}
BILLING_ORPHAN_FIELDS = ["billing_customer_name",
                         "billing_ic_number", "billing_address"]

INSTALL_LABEL_MAP = {
    "Contact Person": "install_contact_person",
    "Tel No (Mobile 1)": "install_mobile1",
    "Tel No (Mobile 2)": "install_mobile2",
    "Tel No (House)": "install_house_phone",
    "Tel No (Office)": "install_office_phone",
    "Email": "install_email",
}
INSTALL_ORPHAN_FIELDS = ["install_address"]

EMERGENCY_LABEL_MAP = {
    "Contact Name": "emergency_contact_name",
    "Contact Number": "emergency_contact_number",
    "Relation": "emergency_relation",
}

HEADER_TEXTS = {
    "Billing Address & Contact",
    "Installation/Service Address & Contact",
    "Emergency Contact",
}

# --- Detail screen: Sales Info tab (confirmed from XML) ---
SALES_INFO_LABEL_MAP = {
    "Sales No.": "sales_info_sales_no",
    "Current Stage": "current_stage",
    "Product": "sales_info_product",
    "OutRight Price": "outright_price",
    "Sales Date": "sales_date",
    "Rental Fees (Monthly)": "rental_fee_monthly",
    "Rental Processing Fees": "rental_processing_fee",
    "Application Type": "application_type",
    "Sales Status": "sales_status",
    "PO Number": "po_number",
    "Promo Code": "promo_code",
    "Rental Scheme": "rental_scheme",
}

OUTPUT_FILE = "cuckoo_export.xlsx"
# Optional: if this .xlsm exists (created once, manually — see the
# write_output()/wa_link_formula() docstrings below for the one-time
# setup), the export writes INTO it instead, preserving its macro so
# the WhatsApp Link column can double-click-open a chat with the full
# message pre-filled, sidestepping HYPERLINK()'s 255-char limit. If it
# doesn't exist, everything still works — the export just falls back
# to the plain .xlsx with a click-to-open-chat-then-paste link instead.
TEMPLATE_FILE = "cuckoo_export_template.xlsm"
OUTPUT_FILE_XLSM = "cuckoo_export.xlsm"

# --- Route planning (address geocoding + grouping) ---
# Runs automatically after scraping, no device/Appium involved — pure
# address lookups against the file, writing into a "Route Plan" sheet
# in the same workbook. Not device-dependent, so it's fine that this
# runs after driver.quit().
GEOCODE_CACHE_FILE = "geocode_cache.json"
# A geocoding match is rejected as "too broad to be useful" if its
# bounding box spans more than this many degrees in either direction
# (roughly 0.05° ≈ 5.5km at Malaysia's latitude). This is what stops a
# detailed address that fails to parse from silently falling back to a
# whole-city or whole-state match.
MAX_MATCH_SPAN_DEGREES = 0.05
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
# Nominatim's usage policy asks for an identifying User-Agent. Feel free
# to edit this to include a real contact if you run this often.
NOMINATIM_USER_AGENT = "cuckoo-plus-route-planner/1.0 (personal workflow tool)"
# Their policy caps usage at 1 request/second — this stays comfortably
# under that.
GEOCODE_REQUEST_DELAY_SECONDS = 1.1
AREA_COLORS = [
    "#1f4e78", "#c00000", "#2e7d32", "#e65100", "#6a1b9a",
    "#00838f", "#ad1457", "#4e342e", "#546e7a", "#9e9d24",
]

WAIT_SECONDS = 10
# raised since smaller, controlled scroll steps need more of them to reach the bottom
MAX_SCROLLS = 300
MAX_STAGNANT_ROUNDS = 2
# fallback step, only used when there's no measured position to target yet (e.g. the very first scroll)
SCROLL_STEP_PERCENT = 0.32
SCROLL_REGION_TOP_FRACTION = 0.40     # stays below the pinned filter header
SCROLL_REGION_HEIGHT_FRACTION = 0.48


# ============================================================
# Low-level helpers
# ============================================================

def build_driver():
    options = UiAutomator2Options()
    options.platform_name = "Android"
    options.app_package = APP_PACKAGE
    options.app_activity = APP_ACTIVITY
    options.no_reset = True
    return webdriver.Remote("http://127.0.0.1:4723", options=options)


def wait_for(driver, selector, timeout=WAIT_SECONDS):
    by, value = selector
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )


def read_text_safe(el):
    try:
        return el.text.strip()
    except Exception:
        return ""


def parse_bounds(bounds_str):
    """'[x1,y1][x2,y2]' -> (x1, y1, x2, y2), or None if unparseable."""
    if not bounds_str:
        return None
    m = re.match(r"\[(\d+),(\d+)\]\[(\d+),(\d+)\]", bounds_str)
    if not m:
        return None
    return tuple(int(g) for g in m.groups())


def tap_element(driver, element):
    """
    Taps by screen coordinates instead of calling .click() directly —
    some Appium-Python-Client / Selenium version combinations throw
    "Wrong parameters applied for elementClick" on UiAutomator2. This
    sidesteps that bug entirely.
    """
    parsed = parse_bounds(element.get_attribute("bounds"))
    if not parsed:
        raise RuntimeError(
            "Could not read bounds — can't compute a tap point for this element.")
    x1, y1, x2, y2 = parsed
    cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
    driver.execute_script("mobile: clickGesture", {"x": cx, "y": cy})


def scroll_down(driver, percent=None):
    """
    Uses "mobile: scrollGesture" rather than "mobile: swipeGesture" —
    swipeGesture performs a FLING, which has momentum: Android keeps
    scrolling after the simulated finger lifts, and how far it travels
    depends on gesture velocity in a way that's hard to predict. That
    inconsistency was very likely why scrolling sometimes jumped past
    several customers at once. scrollGesture instead moves a fixed,
    controlled fraction of the given area with no fling.

    `percent`, when given, comes from compute_scroll_percent() — a
    measured amount based on exactly where the last already-processed
    customer's card sits on screen right now, rather than a guessed
    fixed distance. Falls back to SCROLL_STEP_PERCENT only when there's
    no measurement to base it on yet (the very first scroll).

    Note: for scrollGesture, "direction" describes which way the
    CONTENT moves (not the simulated finger) — "down" reveals further/
    later items in the list, which is what swipeGesture called "up".
    """
    if percent is None:
        percent = SCROLL_STEP_PERCENT
    size = driver.get_window_size()
    width, height = size["width"], size["height"]
    driver.execute_script("mobile: scrollGesture", {
        "left": int(width * 0.1),
        "top": int(height * SCROLL_REGION_TOP_FRACTION),
        "width": int(width * 0.8),
        "height": int(height * SCROLL_REGION_HEIGHT_FRACTION),
        "direction": "down",
        "percent": percent,
    })
    time.sleep(0.6)
    dismiss_keyboard_if_present(driver)


def compute_scroll_percent(driver, target_top_y, margin_fraction=0.04):
    """
    Computes the exact scroll fraction needed to bring `target_top_y`
    (a real, measured card position) up near the top of the scroll
    region — "scroll to the line separating this card from the next,"
    rather than a blind fixed distance.

    A small margin is subtracted so the target lands just inside the
    visible region instead of exactly on the edge, avoiding the
    partially-rendered-row problem.
    """
    size = driver.get_window_size()
    height = size["height"]
    region_top = height * SCROLL_REGION_TOP_FRACTION
    region_height = height * SCROLL_REGION_HEIGHT_FRACTION

    desired_shift = (target_top_y - region_top) - \
        (margin_fraction * region_height)
    percent = desired_shift / region_height

    # Guardrails: never scroll by ~nothing (no progress) or overshoot
    # past the measured area (which would defeat the point of measuring).
    return max(0.08, min(0.95, percent))


def dismiss_keyboard_if_present(driver):
    """Defensive safety net — closes the keyboard if anything ever
    accidentally focuses a text field again, so it doesn't silently
    corrupt the next read."""
    try:
        driver.execute_script("mobile: hideKeyboard")
    except Exception:
        pass  # no keyboard was showing, or the command isn't supported — fine either way


def pair_fields(elements, known_labels=None, skip_texts=None, y_tol=8):
    """
    Generic label/value pairing based on RELATIVE position, not fixed
    pixel coordinates — this makes it work across different screen
    resolutions/devices, since it never assumes a label sits at a
    specific x value. Elements are grouped into rows by shared
    y-coordinate; within each row, the first element (left to right)
    whose text matches a KNOWN label is treated as the label, and the
    element immediately after it becomes its value. Anything sitting
    further left than the label (e.g. a small numbered badge next to
    each card) is simply ignored rather than mistaken for the label.

    Returns (row, orphans):
      row     -> {label_text: value_text} for every recognized label
      orphans -> list of value texts (top-to-bottom order) for rows
                 that don't match a known label — i.e. unlabeled
                 fields like a bare address or name with no caption.
    """
    skip_texts = skip_texts or set()
    entries = []
    for el in elements:
        parsed = parse_bounds(el.get_attribute("bounds"))
        if not parsed:
            continue
        x1, y1, _, _ = parsed
        text = read_text_safe(el)
        if text in skip_texts:
            continue
        entries.append((y1, x1, text))

    entries.sort(key=lambda e: (e[0], e[1]))

    rows = []
    for entry in entries:
        placed = False
        for row in rows:
            if abs(row[0][0] - entry[0]) <= y_tol:
                row.append(entry)
                placed = True
                break
        if not placed:
            rows.append([entry])

    row_dict = {}
    orphans = []
    for row in rows:
        row_sorted = sorted(row, key=lambda e: e[1])  # left to right
        row_y = row_sorted[0][0]

        if known_labels is not None:
            # Find the first element in this row whose text is a genuine
            # known label — this way something sitting further left (like
            # a small numbered badge next to each card) gets ignored
            # instead of being mistaken for the label itself.
            label_idx = next((i for i, e in enumerate(
                row_sorted) if e[2] in known_labels), None)
            if label_idx is not None:
                label_text = row_sorted[label_idx][2]
                value_text = row_sorted[label_idx +
                                        1][2] if label_idx + 1 < len(row_sorted) else ""
                row_dict[label_text] = value_text
                # anything before label_idx (e.g. a badge number) is just ignored
            else:
                for _, _, text in row_sorted:
                    orphans.append((row_y, text))
        else:
            if len(row_sorted) >= 2:
                row_dict[row_sorted[0][2]] = row_sorted[-1][2]
            else:
                orphans.append((row_y, row_sorted[0][2]))

    orphans.sort(key=lambda t: t[0])
    return row_dict, [t for _, t in orphans]


# ============================================================
# List screen
# ============================================================

def group_into_rows(entries, y_tol=8):
    """entries: list of (y1, x1, text). Returns rows: list of rows,
    each row a list of (y1, x1, text) sorted left-to-right, rows
    themselves sorted top-to-bottom."""
    entries = sorted(entries, key=lambda e: (e[0], e[1]))
    rows = []
    for e in entries:
        placed = False
        for row in rows:
            if abs(row[0][0] - e[0]) <= y_tol:
                row.append(e)
                placed = True
                break
        if not placed:
            rows.append([e])
    rows = [sorted(r, key=lambda e: e[1]) for r in rows]
    rows.sort(key=lambda r: r[0][0])
    return rows


def rows_to_fields(rows_chunk, known_labels):
    """Same label-search-per-row logic as pair_fields, applied to an
    already-sliced chunk of rows belonging to one customer."""
    result = {}
    for row in rows_chunk:
        label_idx = next((i for i, e in enumerate(
            row) if e[2] in known_labels), None)
        if label_idx is not None:
            label_text = row[label_idx][2]
            value_text = row[label_idx + 1][2] if label_idx + \
                1 < len(row) else ""
            result[label_text] = value_text
    return result


def get_visible_customers(driver):
    """
    Reads EVERY TextView and Button on screen in one global query (no
    per-card scoped searches — those don't reliably scope on this
    Appium/UiAutomator2 setup, which is what broke the earlier version),
    then figures out which elements belong to which customer purely by
    on-screen position: each "NS No" row marks where a new card starts.

    Returns a list of {"row": {...parsed fields...}, "button": element_or_None}
    for every customer currently visible.
    """
    wait_for(driver, (AppiumBy.XPATH,
             '//android.widget.TextView[@text="NS No"]'))

    text_elements = driver.find_elements(
        AppiumBy.CLASS_NAME, "android.widget.TextView")
    entries = []
    for el in text_elements:
        parsed = parse_bounds(el.get_attribute("bounds"))
        if not parsed:
            continue
        x1, y1, _, _ = parsed
        entries.append((y1, x1, read_text_safe(el)))
    rows = group_into_rows(entries)

    marker_indices = [i for i, row in enumerate(
        rows) if any(t == "NS No" for _, _, t in row)]
    if DEBUG:
        print(
            f"  [debug] found {len(rows)} row(s) total, {len(marker_indices)} 'NS No' marker(s)")

    buttons = driver.find_elements(AppiumBy.XPATH, '//android.widget.Button')
    button_positions = []
    for b in buttons:
        parsed = parse_bounds(b.get_attribute("bounds"))
        if parsed:
            button_positions.append((parsed[1], b))
    button_positions.sort(key=lambda t: t[0])

    known_labels = set(LABEL_FIELD_MAP.keys())
    customers = []
    for idx, start in enumerate(marker_indices):
        end = marker_indices[idx + 1] if idx + \
            1 < len(marker_indices) else len(rows)
        chunk = rows[start:end]
        card_top_y = chunk[0][0][0]
        next_top_y = rows[marker_indices[idx + 1]][0][0] if idx + \
            1 < len(marker_indices) else float("inf")

        field_dict = rows_to_fields(chunk, known_labels)
        parsed_row = {LABEL_FIELD_MAP[k]: v for k,
                      v in field_dict.items() if k in LABEL_FIELD_MAP}

        matching_button = next(
            (b_el for b_y, b_el in button_positions if card_top_y <= b_y < next_top_y), None)
        # A card missing any expected field (most commonly Cust Name,
        # since it's the LAST field on the card) usually means it's only
        # partially scrolled into view — its bottom hasn't fully rendered
        # yet, not that the data is genuinely blank.
        complete = len(field_dict) == len(LABEL_FIELD_MAP)
        customers.append({
            "row": parsed_row,
            "button": matching_button,
            "card_top_y": card_top_y,
            "complete": complete,
        })

    if DEBUG:
        for c in customers:
            print(
                f"  [debug] parsed customer: {c['row']}  (button found: {c['button'] is not None})")

    return customers


def get_visible_customers_stable(driver, max_attempts=4, settle_delay=0.4):
    """
    Reads the screen repeatedly until two consecutive reads agree on
    which NS numbers are visible. A single read can catch the view
    mid-render — Android list rows get REUSED as you scroll, so reading
    too early can show a row still holding the PREVIOUS customer's name
    while its NS No has already updated to the new one. That's what was
    causing blank names and, worse, a name attached to the wrong NS
    number. Waiting for two matching reads in a row avoids trusting a
    transitional, half-updated state.
    """
    prev_signature = None
    customers = []
    for _ in range(max_attempts):
        customers = get_visible_customers(driver)
        signature = tuple(c["row"].get(KEY_FIELD, "") for c in customers)
        if signature == prev_signature and signature:
            return customers
        prev_signature = signature
        time.sleep(settle_delay)
    return customers


# ============================================================
# Popup menu ("Choose Option") -> detail screen
# ============================================================

def select_popup_option(driver, option_text, timeout=WAIT_SECONDS):
    # normalize-space() handles the extra leading spaces the app puts in these labels
    xpath = f'//android.widget.TextView[normalize-space(@text)="{option_text}"]'
    el = wait_for(driver, (AppiumBy.XPATH, xpath), timeout=timeout)
    tap_element(driver, el)


# ============================================================
# Detail screen: Address/Contact Info tab
# ============================================================

def get_billing_elements(driver):
    wait_for(driver, (AppiumBy.XPATH,
             '//android.widget.TextView[@text="Billing Address & Contact"]'))
    return driver.find_elements(
        AppiumBy.XPATH,
        '//android.widget.TextView[@text="Billing Address & Contact"]'
        '/following-sibling::android.view.ViewGroup[1]//android.widget.TextView'
    )


def get_installation_elements(driver):
    wait_for(driver, (AppiumBy.XPATH,
             '//android.widget.TextView[@text="Installation/Service Address & Contact"]'))
    return driver.find_elements(
        AppiumBy.XPATH,
        '//android.widget.TextView[@text="Installation/Service Address & Contact"]'
        '/parent::android.view.ViewGroup//android.widget.TextView'
    )


def get_emergency_elements(driver):
    wait_for(driver, (AppiumBy.XPATH,
             '//android.widget.TextView[@text="Emergency Contact"]'))
    return driver.find_elements(
        AppiumBy.XPATH,
        '//android.widget.TextView[@text="Emergency Contact"]'
        '/parent::android.view.ViewGroup//android.widget.TextView'
    )


def get_sales_info_elements(driver):
    wait_for(driver, (AppiumBy.XPATH,
             '//android.widget.TextView[@text="Current Stage"]'), timeout=8)
    return driver.find_elements(AppiumBy.XPATH, '//android.widget.TextView')


def read_full_detail(driver):
    record = {}

    # --- Address/Contact Info tab (shown by default) ---
    billing_row, billing_orphans = pair_fields(
        get_billing_elements(driver),
        known_labels=set(BILLING_LABEL_MAP.keys()), skip_texts=HEADER_TEXTS
    )
    for label_text, field_name in BILLING_LABEL_MAP.items():
        record[field_name] = billing_row.get(label_text, "")
    for i, field_name in enumerate(BILLING_ORPHAN_FIELDS):
        record[field_name] = billing_orphans[i] if i < len(
            billing_orphans) else ""

    install_row, install_orphans = pair_fields(
        get_installation_elements(driver),
        known_labels=set(INSTALL_LABEL_MAP.keys()), skip_texts=HEADER_TEXTS
    )
    for label_text, field_name in INSTALL_LABEL_MAP.items():
        record[field_name] = install_row.get(label_text, "")
    for i, field_name in enumerate(INSTALL_ORPHAN_FIELDS):
        record[field_name] = install_orphans[i] if i < len(
            install_orphans) else ""

    emergency_row, _ = pair_fields(
        get_emergency_elements(driver),
        known_labels=set(EMERGENCY_LABEL_MAP.keys()), skip_texts=HEADER_TEXTS
    )
    for label_text, field_name in EMERGENCY_LABEL_MAP.items():
        record[field_name] = emergency_row.get(label_text, "")

    # --- Switch to Sales Info tab ---
    sales_tab = wait_for(
        driver, (AppiumBy.XPATH, '//android.widget.TextView[@text="Sales Info"]'))
    tap_element(driver, sales_tab)
    # get_sales_info_elements() below already waits/polls for "Current Stage"
    # to appear, so no fixed sleep needed here.

    sales_row, _ = pair_fields(
        get_sales_info_elements(driver),
        known_labels=set(SALES_INFO_LABEL_MAP.keys())
    )
    for label_text, field_name in SALES_INFO_LABEL_MAP.items():
        record[field_name] = sales_row.get(label_text, "")

    return record


# ============================================================
# Detail screen: CCS Note popup (filter/consumable change data)
# ============================================================
#
# A scrollable list of "cards" — one per physical filter/consumable UNIT
# installed, which is why the exact same product can appear several
# times in a row (e.g. "FT-1001 Sediment Filter 8 Inch" showing up 6
# times for one customer, because they have 6 identical units). Each
# full card has: a product name, a small per-unit number (1, 2, 3...),
# an interval like "(4 months)", "Last Change" -> a date, "Next Change"
# -> a date. Only product name + Last Change date are actually kept —
# interval and Next Change aren't part of what's needed here.
#
# A card is only kept if BOTH of these hold:
#   - its product name is real text, not a bare number and not the
#     literal label "Last Change"/"Next Change"
#   - it has a non-empty Last Change date
# Both checks exist because of a real, confirmed pattern in captured
# data: alongside each genuine card, the same screen sometimes also
# yields a partial/garbled read of it — either the standalone "Next
# Change" label picked up on its own, or the small per-unit number
# landing where the product name should be — and in every observed
# case, that garbled read has a BLANK Last Change while the genuine
# card next to it doesn't. Filtering on "has a real name AND has a
# Last Change" reliably keeps the real entries and drops the artifacts.
#
# "Unique" cards: product name + Last Change date must both match for
# two cards to be treated as duplicates and collapsed into one
# (CCS_DEDUP_FIELDS below). The small per-unit number is deliberately
# not part of that — this is what collapses several near-identical
# "FT-1001 Sediment Filter 8 Inch" cards, differing only by unit
# number, into a single row.

CCS_DEDUP_FIELDS = ("product", "last_change")
# Safety cap on scrolling WITHIN one customer's CCS Note screen — same
# spirit as MAX_SCROLLS above, just a separate, smaller budget since
# this screen has far fewer cards than the full customer list does.
MAX_CCS_CARD_SCROLLS = 30

_CCS_RESERVED_LABELS = {"Last Change", "Next Change"}


def _is_valid_ccs_product_name(product):
    if not product:
        return False
    if product in _CCS_RESERVED_LABELS:
        return False
    if product.strip().isdigit():
        return False
    return True


def get_ccs_note_cards(driver):
    """Reads every currently-visible card on the CCS Note screen,
    keeping only ones that pass the validity checks above."""
    wait_for(driver, (AppiumBy.XPATH,
             '//android.widget.TextView[@text="CCS Note"]'))

    card_elements = driver.find_elements(
        AppiumBy.XPATH,
        '//androidx.viewpager.widget.ViewPager//android.view.ViewGroup[@clickable="true"]'
    )

    cards = []
    for card_el in card_elements:
        parsed = parse_bounds(card_el.get_attribute("bounds"))
        if not parsed:
            continue
        text_elements = card_el.find_elements(
            AppiumBy.CLASS_NAME, "android.widget.TextView")
        texts = [read_text_safe(t) for t in text_elements]
        texts = [t for t in texts if t]  # drop empty separator TextViews

        if not texts:
            continue

        product = texts[0]
        last_change = ""
        for i, t in enumerate(texts):
            if t == "Last Change" and i + 1 < len(texts):
                last_change = texts[i + 1]

        if not _is_valid_ccs_product_name(product):
            continue

        cards.append({
            "product": product,
            "last_change": last_change,
            "card_top_y": parsed[1],
        })

    return cards


def get_all_ccs_note_cards(driver):
    """
    Scrolls through the whole CCS Note screen, collecting cards as it
    goes. Deduplicates along the way using CCS_DEDUP_FIELDS — this
    collapses both genuinely repeated cards (same product/last change,
    different unit number) AND the same card being seen twice after a
    small scroll, with the same check. Finishes with a reconciliation
    pass (see _reconcile_ccs_cards) that separates real "not yet
    serviced" entries from parsing artifacts.
    """
    seen_keys = set()
    unique_cards = []
    scroll_count = 0
    stagnant_rounds = 0

    while scroll_count < MAX_CCS_CARD_SCROLLS and stagnant_rounds < MAX_STAGNANT_ROUNDS:
        cards = get_ccs_note_cards(driver)
        if not cards:
            break

        new_this_round = 0
        for card in cards:
            key = tuple(card[f] for f in CCS_DEDUP_FIELDS)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            unique_cards.append(card)
            new_this_round += 1

        stagnant_rounds = 0 if new_this_round > 0 else stagnant_rounds + 1

        target_percent = compute_scroll_percent(
            driver, cards[-1]["card_top_y"])
        scroll_down(driver, percent=target_percent)
        scroll_count += 1

    return _reconcile_ccs_cards(unique_cards)


def _reconcile_ccs_cards(cards):
    """
    Separates real "not yet serviced" entries from parsing artifacts —
    both look identical in isolation (same product, blank Last Change),
    so this has to look at each PRODUCT's entries together to tell them
    apart:

      - If a product has at least one entry WITH a Last Change date,
        any blank-date entries for that same product are almost
        certainly a partial/duplicate read of that same card (a
        confirmed real pattern — see get_ccs_note_cards' docstring),
        so they're dropped. Every distinct dated entry is kept (a
        product can legitimately have more than one physical unit,
        serviced on different dates).
      - If a product has NO dated entry at all, it's kept as-is with a
        blank date — that's a genuine filter that just hasn't had its
        first change recorded yet, not an artifact.
    """
    by_product = {}
    for card in cards:
        by_product.setdefault(card["product"], []).append(card)

    reconciled = []
    for product, product_cards in by_product.items():
        dated = [c for c in product_cards if c["last_change"]]
        if dated:
            reconciled.extend(dated)
        else:
            reconciled.append(product_cards[0])
    return reconciled


# ============================================================
# Main loop: scroll + scrape until nothing new appears
# ============================================================

def run():
    driver = build_driver()
    time.sleep(3)

    all_records = []
    all_ccs_rows = []
    seen_keys = set()
    stagnant_rounds = 0
    scroll_count = 0

    try:
        while True:
            if LIST_LIMIT and len(all_records) >= LIST_LIMIT:
                print(f"Reached LIST_LIMIT of {LIST_LIMIT} — stopping.")
                break

            customers = get_visible_customers_stable(driver)

            next_customer = None
            partial_customer = None
            for c in customers:
                key = c["row"].get(KEY_FIELD)
                if not key or key in seen_keys:
                    continue
                if c["button"] is None or not c.get("complete", True):
                    # Card is likely only partially scrolled into view — its
                    # NS No rendered (enough to be found and keyed), but its
                    # button and/or its later fields (Cust Name is the last
                    # field on the card, so it's usually the first casualty)
                    # haven't fully rendered yet. Don't mark it seen; remember
                    # it so we can scroll IT specifically into full view,
                    # rather than treating this like "nothing new at all."
                    if partial_customer is None:
                        partial_customer = c
                    continue
                next_customer = c
                break

            if next_customer is None:
                if partial_customer is not None:
                    # There IS a new customer here — it's just not fully
                    # rendered yet. Scroll targeted at THIS card's own
                    # position (not the generic "last customer" case below)
                    # so it comes fully into view instead of being skipped.
                    scroll_count += 1
                    if scroll_count >= MAX_SCROLLS:
                        print(
                            "Hit the safety scroll limit — stopping to avoid an infinite loop.")
                        break
                    target_percent = compute_scroll_percent(
                        driver, partial_customer["card_top_y"])
                    if DEBUG:
                        print(f"  [debug] {partial_customer['row'].get(KEY_FIELD)} not fully rendered yet — "
                              f"scrolling it into view (percent={target_percent:.3f})")
                    scroll_down(driver, percent=target_percent)
                    continue

                stagnant_rounds += 1
                if stagnant_rounds >= MAX_STAGNANT_ROUNDS:
                    print(
                        "No new customers found after scrolling — reached the end of the list.")
                    break
                scroll_count += 1
                if scroll_count >= MAX_SCROLLS:
                    print(
                        "Hit the safety scroll limit — stopping to avoid an infinite loop.")
                    break
                target_percent = compute_scroll_percent(
                    driver, customers[-1]["card_top_y"]) if customers else SCROLL_STEP_PERCENT
                if DEBUG:
                    print(
                        f"  [debug] scrolling by measured percent={target_percent:.3f}")
                scroll_down(driver, percent=target_percent)
                continue

            stagnant_rounds = 0
            next_row = next_customer["row"]
            key = next_row[KEY_FIELD]
            seen_keys.add(key)
            print(
                f"[{len(all_records) + 1}] Opening: {key} ({next_row.get('cust_name', '')})")

            try:
                button = next_customer["button"]
                if button is None:
                    raise RuntimeError(
                        "No 'View Order' button found near this customer's row")

                # --- Visit 1: View Order (sales/install details) ---
                try:
                    try:
                        tap_element(driver, button)
                    except Exception as e:
                        raise RuntimeError(
                            f"[stage: tapping row button] {type(e).__name__}: {e}")

                    try:
                        # select_popup_option() already waits/polls for the popup
                        # to appear — no fixed sleep needed before it.
                        select_popup_option(driver, "View Order")
                    except Exception as e:
                        raise RuntimeError(
                            f"[stage: selecting 'View Order' from popup] {type(e).__name__}: {e}")

                    try:
                        wait_for(
                            driver, (AppiumBy.XPATH, '//android.widget.Button[@text="Customer Information"]'))
                    except Exception as e:
                        raise RuntimeError(
                            f"[stage: waiting for Customer Information screen] {type(e).__name__}: {e}")

                    try:
                        detail = read_full_detail(driver)
                    except Exception as e:
                        raise RuntimeError(
                            f"[stage: reading detail screen fields] {type(e).__name__}: {e}")

                    all_records.append({**next_row, **detail})
                except Exception as e:
                    print(f"  !! Skipped View Order for {key}: {e}")
                finally:
                    driver.back()
                    # get_visible_customers() below already waits/polls for "NS
                    # No" to reappear, so no sleep needed here.

                # --- Visit 2: CCS Note (filter/consumable change data) ---
                # A fresh element lookup is required here — the `button`
                # WebElement from before driver.back() is stale now (the
                # underlying UI tree changed), so it can't just be reused for
                # a second tap the way it could within a single visit.
                try:
                    try:
                        customers_again = get_visible_customers_stable(driver)
                        this_customer_again = next(
                            (c for c in customers_again if c["row"].get(KEY_FIELD) == key), None)
                        if this_customer_again is None or this_customer_again["button"] is None:
                            raise RuntimeError(
                                "Could not re-find this customer's row for CCS Note")
                    except Exception as e:
                        raise RuntimeError(
                            f"[stage: re-finding row after View Order] {type(e).__name__}: {e}")

                    try:
                        tap_element(driver, this_customer_again["button"])
                    except Exception as e:
                        raise RuntimeError(
                            f"[stage: tapping row button] {type(e).__name__}: {e}")

                    try:
                        select_popup_option(driver, "CCS Note")
                    except Exception as e:
                        raise RuntimeError(
                            f"[stage: selecting 'CCS Note' from popup] {type(e).__name__}: {e}")

                    try:
                        cards = get_all_ccs_note_cards(driver)
                    except Exception as e:
                        raise RuntimeError(
                            f"[stage: reading CCS Note cards] {type(e).__name__}: {e}")

                    for card in cards:
                        all_ccs_rows.append({
                            "sales_no": next_row.get("sales_no", ""),
                            "cust_name": next_row.get("cust_name", ""),
                            "product": card["product"],
                            "last_change": card["last_change"],
                        })
                except Exception as e:
                    print(f"  !! Skipped CCS Note for {key}: {e}")
                finally:
                    driver.back()
            except Exception as e:
                print(f"  !! Skipped {key} entirely due to error: {e}")

    finally:
        driver.quit()

    written_to = write_output(all_records, all_ccs_rows)
    print(f"Done. Wrote {len(all_records)} record(s) and {len(all_ccs_rows)} "
          f"CCS Note row(s) to {written_to}")


# ============================================================
# Export column order:
#   A=sales_no, B=appt_date, C=install_contact_person,
#   D=install_mobile1, E=install_address, F=proposed_date (typed by
#   hand), G=WhatsApp Link, H=sales_info_product, I=Filters (from CCS
#   Note data), J=WhatsApp Message, K=wa_number (hidden helper).
# ============================================================

ALL_COLUMNS = [
    "sales_no", "appt_date", "install_contact_person", "install_mobile1",
    "install_address", "proposed_date", "WhatsApp Link", "sales_info_product",
    "Filters", "WhatsApp Message", "wa_number",
]


def clean_phone_for_wa(raw):
    """
    Strips everything except digits, then converts a local Malaysian
    mobile number (leading 0, e.g. "012-345 6789") into the international
    format wa.me links need (leading 60, no separators, e.g.
    "60123456789"). A number that's already in some other international
    format (doesn't start with 0 after stripping) is left as digits-only
    — we can't safely guess a country code that isn't already there.
    Returns "" if there's nothing usable.
    """
    digits = re.sub(r"\D", "", raw or "")
    if not digits:
        return ""
    if digits.startswith("0"):
        digits = "60" + digits[1:]
    return digits


def message_formula(row):
    """
    Excel formula for one row's WhatsApp Message cell, matching the
    copywriting/formatting shown in the reference screenshot. Single
    asterisks are WhatsApp's own bold syntax, not markdown. Column
    letters below map to the export column order above: A=sales_no,
    E=install_address, F=proposed_date, H=sales_info_product.

    Note: if you copy this CELL (Ctrl+C, not double-click) and paste
    into something like Notepad, you'll see the whole value wrapped in
    quotes. That's Excel's own clipboard behavior (CSV-style quoting)
    kicking in because the text contains commas and line breaks — it's
    not part of the cell's actual value or this formula, and it doesn't
    matter once you're using the WhatsApp Link column instead of
    copying this text by hand.
    """
    tarikh_part = (
        f'IFERROR(TEXT(F{row},"DD/MM/YYYY") & " (" & '
        f'CHOOSE(WEEKDAY(F{row},2),"Isnin","Selasa","Rabu","Khamis","Jumaat","Sabtu","Ahad") & ")", F{row})'
    )
    return (
        f'=IF(F{row}="","","Selamat sejahtera Tuan/Puan," & CHAR(10) & CHAR(10) & '
        f'"Saya *Hanis*, CUCKOO+ Service Specialist (NDS35095). Saya memohon maaf jika saya menghubungi anda pada waktu yang tidak sesuai." & CHAR(10) & CHAR(10) & '
        f'"Saya ingin mengesahkan jika saya boleh membuat lawatan servis seperti di bawah." & CHAR(10) & CHAR(10) & '
        f'"*Tarikh:* " & {tarikh_part} & CHAR(10) & '
        f'"*Alamat:* " & E{row} & CHAR(10) & '
        f'"*Produk:* " & H{row} & CHAR(10) & '
        f'"*Nombor Pesanan:* " & A{row} & CHAR(10) & CHAR(10) & '
        f'"Terima kasih, sokongan dan kerjasama Tuan/Puan amat saya hargai.")'
    )


def wa_link_formula(row, phone_digits):
    """
    Excel formula for one row's WhatsApp Link cell.

    IMPORTANT LIMITATION: this can't pre-fill the message the way a
    wa.me "?text=" link normally would. Excel's own HYPERLINK() worksheet
    function hard-caps its link_location argument at 255 characters —
    if it's longer, Excel returns #VALUE! instead of opening anything.
    The encoded message (greeting + explanation + address etc.) is
    always well past that, no matter how it's trimmed, so a formula-
    based link genuinely cannot carry the pre-filled text.

    What this DOES do: opens the right WhatsApp chat directly (a plain
    "https://wa.me/<number>" link, comfortably under 255 chars), so you
    only need to copy the WhatsApp Message cell (column J) and paste it
    in — no manual number lookup or searching for the contact. Column
    letters map to the export column order above: F=proposed_date.
    """
    if not phone_digits:
        return '="No phone number found"'
    return (
        f'=IF(F{row}="","",HYPERLINK('
        f'"https://wa.me/{phone_digits}",'
        f'"Open WhatsApp Chat"))'
    )


def _populate_sheet(ws, records, filters_by_sales_no):
    """
    Fills in headers + rows on an already-created worksheet (either a
    fresh one, or the one already inside the macro-enabled template).
    Shared by both output paths in write_output() below so the two
    stay in sync automatically. filters_by_sales_no maps sales_no ->
    the combined multi-line Filters text (see _build_filters_lookup).
    """
    FONT = "Arial"
    header_fill = PatternFill("solid", fgColor="1F4E78")

    for col_idx, header in enumerate(ALL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center")
    ws["F1"].comment = Comment(
        "Type a date here as DD/MM/YYYY (e.g. 08/08/2026).\n"
        "The WhatsApp Message and WhatsApp Link columns fill themselves in automatically.",
        "cuckoo_scraper.py"
    )
    ws["G1"].comment = Comment(
        "Click to open the right WhatsApp chat directly, then copy the\n"
        "message from column J and paste it in. (Excel's HYPERLINK function\n"
        "can't carry a pre-filled message this long — it caps links at 255\n"
        "characters — so this gets you to the chat, but the message still\n"
        "needs one paste.)\n\n"
        "If this file has the WhatsApp macro installed (see write_output()'s\n"
        "docstring in cuckoo_scraper.py), double-click instead of single-\n"
        "clicking — that opens the chat with the message already filled in,\n"
        "no paste needed.",
        "cuckoo_scraper.py"
    )
    ws["K1"].comment = Comment(
        "Internal use only — the macro reads this to build the full "
        "pre-filled WhatsApp link. Don't edit or delete this column.",
        "cuckoo_scraper.py"
    )

    for row_idx, record in enumerate(records, start=2):
        # A-E: plain scraped fields, in the specified order.
        for col_idx, field in enumerate(
            ["sales_no", "appt_date", "install_contact_person",
             "install_mobile1", "install_address"], start=1
        ):
            ws.cell(row=row_idx, column=col_idx,
                    value=record.get(field, "")).font = Font(name=FONT, size=10)

        # F: proposed_date — typed by hand, no special fill (styled the
        # same as every other cell, on request).
        ws.cell(row=row_idx, column=6,
                value=record.get("proposed_date")).font = Font(name=FONT, size=10)

        phone_digits = clean_phone_for_wa(record.get("install_mobile1", ""))

        # G: WhatsApp Link (formula)
        g_cell = ws.cell(row=row_idx, column=7,
                         value=wa_link_formula(row_idx, phone_digits))
        g_cell.font = Font(name=FONT, size=10,
                           color="1155CC", underline="single")

        # H: sales_info_product
        ws.cell(row=row_idx, column=8,
                value=record.get("sales_info_product", "")).font = Font(name=FONT, size=10)

        # I: Filters — combined text from CCS Note data, looked up by
        # sales_no; blank if this customer had no CCS Note data captured.
        filters_cell = ws.cell(row=row_idx, column=9,
                               value=filters_by_sales_no.get(record.get("sales_no", ""), ""))
        filters_cell.font = Font(name=FONT, size=10)
        filters_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # J: WhatsApp Message (formula)
        j_cell = ws.cell(row=row_idx, column=10,
                         value=message_formula(row_idx))
        j_cell.font = Font(name=FONT, size=10)
        j_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # K: wa_number — hidden helper column, the macro (if installed)
        # reads this directly instead of re-deriving the phone number.
        k_cell = ws.cell(row=row_idx, column=11, value=phone_digits)
        k_cell.font = Font(name=FONT, size=10)

    widths = {"A": 14, "B": 12, "C": 26, "D": 16, "E": 40, "F": 14,
              "G": 18, "H": 16, "I": 40, "J": 60, "K": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["K"].hidden = True
    ws.freeze_panes = "A2"


def _build_filters_lookup(ccs_rows):
    """
    Turns the flat list of {sales_no, cust_name, product, last_change}
    rows from get_all_ccs_note_cards() into {sales_no: combined_text},
    one "Product: date" line per filter (real line breaks, so it reads
    like pressing Alt+Enter between each one), sorted alphabetically by
    product name for a consistent read.
    """
    by_customer = {}
    for r in ccs_rows:
        by_customer.setdefault(r["sales_no"], []).append(
            (r["product"], r["last_change"]))

    lookup = {}
    for sales_no, filters in by_customer.items():
        filters_sorted = sorted(filters, key=lambda f: f[0])
        lookup[sales_no] = "\n".join(
            f"{product}: {last_change if last_change else 'not yet changed'}"
            for product, last_change in filters_sorted
        )
    return lookup


# ============================================================
# Route planning: address geocoding + grouping (Route Plan sheet)
# ============================================================
#
# Runs automatically after scraping finishes — no device/Appium
# involved, this is pure address lookups against OpenStreetMap's free
# Nominatim service, safe to run after driver.quit(). Builds directly
# from the in-memory `records` list (not by re-reading the saved file's
# cell positions — that would be fragile against future column
# reorders, which is exactly what broke the old standalone
# route_planner.py script's approach).
#
#   1. Geocodes each customer's install_address (cached locally in
#      GEOCODE_CACHE_FILE, so re-running doesn't re-geocode addresses
#      already resolved).
#   2. Groups customers by Malaysia's actual administrative hierarchy —
#      state (negeri) > district (daerah) > city/town > suburb/precinct
#      > neighbourhood — pulled from OpenStreetMap's own structured
#      address breakdown, rather than an arbitrary distance radius.
#      This is what correctly separates two addresses that happen to be
#      physically close but are in different named areas (e.g. same
#      precinct, different named sub-development).
#   3. Classifies each address as "Landed" or "High-Rise" from the
#      address text itself.
#   4. Writes a "Route Plan" sheet — ONE sorted view: grouped by Area,
#      and within each area split into a Landed block and a High-Rise
#      block, each listing its stops with a Google Maps link.
#
# ADDRESSES THAT DON'T GEOCODE: some (especially very detailed ones —
# specific block/unit numbers) won't resolve on the first try. This
# automatically retries using a simplified, street/precinct-level
# version of the address if the full one fails. If that still fails,
# the customer is listed separately at the bottom of the Route Plan
# sheet under "Could not place automatically".
#
# NOMINATIM USAGE NOTE: their usage policy asks for a maximum of 1
# request/second and an identifying User-Agent — both handled via
# NOMINATIM_USER_AGENT / GEOCODE_REQUEST_DELAY_SECONDS in the CONFIG
# section above.

def load_geocode_cache():
    if not os.path.exists(GEOCODE_CACHE_FILE):
        return {}
    try:
        with open(GEOCODE_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_geocode_cache(cache):
    with open(GEOCODE_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def _nominatim_query(params):
    """One rate-limited request to Nominatim. Returns the raw first
    result dict (with lat/lon/boundingbox/address/etc.) or None."""
    params = {**params, "addressdetails": 1}
    url = NOMINATIM_URL + "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(
        url, headers={"User-Agent": NOMINATIM_USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            results = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"    [geocode] request failed: {type(e).__name__}: {e}")
        return None
    if not results:
        return None
    return results[0]


# Each entry: (canonical level name, candidate OSM/Nominatim field names
# to check, in order). Malaysia's real admin hierarchy is Negeri (state)
# > Daerah (district) > Bandar/Pekan (city/town) > Presint/Taman/Mukim
# (suburb) > a finer named pocket within that (neighbourhood) — but OSM
# contributors don't always tag things under the exact same field name,
# so each level checks a few plausible alternates.
_HIERARCHY_LEVELS = [
    ("state", ["state"]),
    ("district", ["state_district", "county"]),
    ("city", ["city", "town", "municipality"]),
    ("suburb", ["suburb", "city_district", "borough"]),
    ("neighbourhood", ["neighbourhood", "quarter", "residential", "hamlet"]),
]


def extract_hierarchy(result):
    """Pulls a (state, district, city, suburb, neighbourhood) tuple out
    of a Nominatim result's structured "address" breakdown."""
    address = result.get("address", {}) if result else {}
    levels = []
    for _, candidates in _HIERARCHY_LEVELS:
        value = None
        for field in candidates:
            if address.get(field):
                value = address[field]
                break
        levels.append(value)
    return tuple(levels)


def hierarchy_label(hierarchy):
    """Human-readable path, e.g. 'Selangor > Petaling > Petaling Jaya > SS2'."""
    parts = [p for p in hierarchy if p]
    return " > ".join(parts) if parts else "Unknown area"


def _is_specific_enough(result):
    """
    Rejects matches that are too broad to be useful — e.g. when a
    detailed unit/block address doesn't parse and Nominatim quietly
    falls back to matching just the city or state name. Left
    unchecked, that gives every address in the same city the SAME
    coordinates, silently producing a garbage grouping instead of an
    honest failure. Uses the result's boundingbox width/height as a
    proxy for how precise the match is.
    """
    bbox = result.get("boundingbox")
    if not bbox:
        return False
    south, north, west, east = (float(x) for x in bbox)
    return (north - south) <= MAX_MATCH_SPAN_DEGREES and (east - west) <= MAX_MATCH_SPAN_DEGREES


def _simplify_to_street_level(address):
    """
    Malaysian residential addresses often lead with a unit/block/lot
    code that essentially never exists in OpenStreetMap's data, which
    often makes the full-address query fail entirely. This strips
    everything before the first recognizable street/area keyword
    (JALAN, PERSIARAN, LORONG, PRESINT, TAMAN, BANDAR, LEBUH,
    LINGKARAN), keeping the part that's actually likely to be mapped.
    Returns None if none of those keywords appear.
    """
    keywords = r"(JALAN|PERSIARAN|LORONG|PRESINT|TAMAN|BANDAR|LEBUH|LINGKARAN)"
    match = re.search(keywords, address, re.IGNORECASE)
    if not match:
        return None
    return address[match.start():]


def geocode_address(address, cache):
    """
    Returns {"lat": ..., "lon": ..., "hierarchy": (...)} or None, using
    the cache first. On a cache miss, tries three queries in order,
    keeping the first that's actually usable: the full address; a
    simplified street/precinct-level version; a structured postcode
    query as a last, deliberately coarser resort (not rejected for
    being "too broad", since it's supposed to be).
    """
    key = address.strip()
    if not key:
        return None

    if key in cache:
        cached = cache[key]
        if cached is None:
            return None
        if isinstance(cached, dict) and "hierarchy" in cached:
            return {**cached, "hierarchy": tuple(cached["hierarchy"])}
        # else: old cache format from before hierarchy tracking existed —
        # fall through and re-geocode.

    time.sleep(GEOCODE_REQUEST_DELAY_SECONDS)
    result = _nominatim_query({
        "q": key, "format": "jsonv2", "countrycodes": "my", "limit": 1,
    })
    if result is not None and not _is_specific_enough(result):
        result = None

    if result is None:
        simplified = _simplify_to_street_level(key)
        if simplified:
            time.sleep(GEOCODE_REQUEST_DELAY_SECONDS)
            result = _nominatim_query({
                "q": simplified, "format": "jsonv2", "countrycodes": "my", "limit": 1,
            })
            if result is not None and not _is_specific_enough(result):
                result = None

    if result is None:
        postcode_match = re.search(r"\b(\d{5})\b", key)
        if postcode_match:
            time.sleep(GEOCODE_REQUEST_DELAY_SECONDS)
            result = _nominatim_query({
                "postalcode": postcode_match.group(1), "country": "Malaysia",
                "format": "jsonv2", "limit": 1,
            })

    if result:
        hierarchy = extract_hierarchy(result)
        geocoded = {"lat": float(result["lat"]), "lon": float(
            result["lon"]), "hierarchy": hierarchy}
        print(
            f"    [geocode] \"{key[:40]}...\" -> {hierarchy_label(hierarchy)}")
        cache[key] = {"lat": geocoded["lat"],
                      "lon": geocoded["lon"], "hierarchy": list(hierarchy)}
    else:
        geocoded = None
        cache[key] = None

    save_geocode_cache(cache)
    return geocoded


def gmaps_single_link(lat, lon):
    return f"https://www.google.com/maps/search/?api=1&query={lat},{lon}"


_HIGH_RISE_KEYWORDS = re.compile(
    r"\b(BLOK|BLOCK|TINGKAT|PANGSAPURI|KONDOMINIUM|CONDOMINIUM|CONDO|"
    r"APARTMENT|FLAT|RESIDENSI|RESIDENCE|SUITES?|MENARA|TOWER|PARCEL)\b",
    re.IGNORECASE,
)
_UNIT_CODE_PATTERN = re.compile(
    r"\b[A-Z]-[A-Z]?\d+[A-Z]?-[A-Z]?\d+\b", re.IGNORECASE)


def classify_housing_type(address):
    """"Landed" or "High-Rise", guessed from the address text alone —
    keyword or block-floor-unit code (A-12-05) means high-rise,
    anything without either signal is assumed landed."""
    if _HIGH_RISE_KEYWORDS.search(address) or _UNIT_CODE_PATTERN.search(address):
        return "High-Rise"
    return "Landed"


_STREET_PATTERN = re.compile(
    r"\b(?:JALAN|JLN|LORONG|PERSIARAN|LEBUH|LINGKARAN)\s+"
    r"[A-Z0-9/.\-]+(?:\s+[A-Z0-9/.\-]+)?",
    re.IGNORECASE,
)


def extract_street(address):
    """Pulls the road name/code directly from the address text — used
    to sort stops WITHIN an area/precinct even when OSM has no data for
    that specific lane (see geocode_address's docstring)."""
    match = _STREET_PATTERN.search(address)
    if not match:
        return ""
    return match.group(0).strip().upper()


def _street_sort_key(street):
    """Normalizes minor spacing (e.g. "P11J" vs "P11 J", the same
    road) so they sort/group together."""
    return street.replace(" ", "")


def build_route_plan_sheet(wb, records):
    """
    Builds the "Route Plan" sheet directly on the given workbook, from
    the same `records` about to be written to the Cuckoo Export sheet
    — not by re-reading cells back out of a saved file, which would be
    fragile against column reorders.
    """
    rows = [{
        "sales_no": r.get("sales_no", ""),
        "contact": r.get("install_contact_person", ""),
        "phone": r.get("install_mobile1", ""),
        "address": r.get("install_address", ""),
    } for r in records if r.get("sales_no") and r.get("install_address")]

    if not rows:
        return

    for r in rows:
        r["housing_type"] = classify_housing_type(r["address"])
        r["street"] = extract_street(r["address"])

    print(f"Geocoding {len(rows)} address(es) for the route plan "
          f"(cached ones are instant, new ones take ~1s each)...")
    cache = load_geocode_cache()
    geocoded, failed = [], []
    for i, r in enumerate(rows, start=1):
        result = geocode_address(r["address"], cache)
        if result:
            r["lat"], r["lon"], r["hierarchy"] = result["lat"], result["lon"], result["hierarchy"]
            geocoded.append(r)
        else:
            failed.append(r)
        print(
            f"  [{i}/{len(rows)}] {r['sales_no']}: {'OK' if result else 'could not place'}")

    if not geocoded and not failed:
        return

    areas = {}
    for r in geocoded:
        areas.setdefault(r["hierarchy"], []).append(r)

    def sort_key(hierarchy):
        return tuple(level or "" for level in hierarchy)
    sorted_hierarchies = sorted(areas.keys(), key=sort_key)

    if "Route Plan" in wb.sheetnames:
        del wb["Route Plan"]
    rp = wb.create_sheet("Route Plan")

    FONT = "Arial"
    subheader_fill = PatternFill("solid", fgColor="D9E1F2")
    r_idx = 1

    def write_stop_block(rows_for_block, start_row):
        row = start_row
        headers = ["Sales No", "Contact", "Phone",
                   "Street", "Address (tap to open)"]
        for col_idx, h in enumerate(headers, start=1):
            c = rp.cell(row=row, column=col_idx, value=h)
            c.font = Font(name=FONT, size=10, bold=True)
            c.fill = subheader_fill
        row += 1
        sorted_rows = sorted(
            rows_for_block, key=lambda r: (_street_sort_key(r["street"]), r["sales_no"]))
        for r in sorted_rows:
            rp.cell(row=row, column=1, value=r["sales_no"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=row, column=2, value=r["contact"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=row, column=3, value=r["phone"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=row, column=4, value=r["street"] or "?").font = Font(
                name=FONT, size=10)
            addr_cell = rp.cell(row=row, column=5, value=r["address"])
            addr_cell.hyperlink = gmaps_single_link(r["lat"], r["lon"])
            addr_cell.font = Font(name=FONT, size=10,
                                  color="1155CC", underline="single")
            addr_cell.alignment = Alignment(wrap_text=True)
            row += 1
        return row

    for area_idx, hierarchy in enumerate(sorted_hierarchies):
        area_rows = areas[hierarchy]
        area_color = AREA_COLORS[area_idx % len(AREA_COLORS)]
        cell = rp.cell(row=r_idx, column=1,
                       value=f"{hierarchy_label(hierarchy)} — {len(area_rows)} address(es)")
        cell.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=area_color.lstrip("#"))
        rp.merge_cells(start_row=r_idx, start_column=1,
                       end_row=r_idx, end_column=5)
        r_idx += 1

        landed = [r for r in area_rows if r["housing_type"] == "Landed"]
        high_rise = [r for r in area_rows if r["housing_type"] == "High-Rise"]
        for label, subset in [("Landed", landed), ("High-Rise", high_rise)]:
            if not subset:
                continue
            sub_cell = rp.cell(row=r_idx, column=1,
                               value=f"{label} ({len(subset)})")
            sub_cell.font = Font(name=FONT, size=11, bold=True, italic=True)
            rp.merge_cells(start_row=r_idx, start_column=1,
                           end_row=r_idx, end_column=5)
            r_idx += 1
            r_idx = write_stop_block(subset, r_idx)
            r_idx += 1

        r_idx += 1

    if failed:
        cell = rp.cell(row=r_idx, column=1,
                       value="Could not place automatically — assign manually")
        cell.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        rp.merge_cells(start_row=r_idx, start_column=1,
                       end_row=r_idx, end_column=5)
        r_idx += 1
        headers = ["Sales No", "Contact", "Phone",
                   "Street", "Address", "Housing Type"]
        for col_idx, h in enumerate(headers, start=1):
            c = rp.cell(row=r_idx, column=col_idx, value=h)
            c.font = Font(name=FONT, size=10, bold=True)
            c.fill = subheader_fill
        r_idx += 1
        for r in sorted(failed, key=lambda r: (_street_sort_key(r["street"]), r["sales_no"])):
            rp.cell(row=r_idx, column=1, value=r["sales_no"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=r_idx, column=2, value=r["contact"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=r_idx, column=3, value=r["phone"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=r_idx, column=4, value=r["street"] or "?").font = Font(
                name=FONT, size=10)
            addr_cell = rp.cell(row=r_idx, column=5, value=r["address"])
            addr_cell.alignment = Alignment(wrap_text=True)
            addr_cell.font = Font(name=FONT, size=10)
            rp.cell(row=r_idx, column=6, value=r["housing_type"]).font = Font(
                name=FONT, size=10)
            r_idx += 1

    rp.column_dimensions["A"].width = 14
    rp.column_dimensions["B"].width = 24
    rp.column_dimensions["C"].width = 16
    rp.column_dimensions["D"].width = 22
    rp.column_dimensions["E"].width = 50
    rp.column_dimensions["F"].width = 14

    print(f"Route plan: {len(geocoded)} address(es) grouped into {len(areas)} area(s), "
          f"{len(failed)} need manual placement.")


def write_output(records, ccs_rows=None):
    """
    Writes the export with the columns in the order specified at the
    top of this section (sales_no, appt_date, install_contact_person,
    install_mobile1, install_address, proposed_date, WhatsApp Link,
    sales_info_product, Filters, WhatsApp Message, wa_number) — all in
    ONE sheet, no separate CCS Notes sheet.

    TWO POSSIBLE OUTPUTS, chosen automatically:

    1. If TEMPLATE_FILE ("cuckoo_export_template.xlsm") exists, this
       writes INTO a copy of it (loaded with keep_vba=True so its
       macro survives) and saves as OUTPUT_FILE_XLSM. In that file,
       double-clicking a WhatsApp Link cell runs the macro, which opens
       WhatsApp with the message already filled in — no paste needed,
       because VBA's FollowHyperlink isn't subject to the 255-character
       cap that the HYPERLINK() *formula* has.

    2. Otherwise, this falls back to a plain OUTPUT_FILE (.xlsx) with
       no macro — the WhatsApp Link column still works via single
       click, it just opens the bare chat (see wa_link_formula()'s
       docstring) rather than pre-filling the message.

    ONE-TIME SETUP for option 1 (only needs doing once, ever —
    openpyxl can't write compiled VBA itself, so this part is manual).
    IMPORTANT: if you already pasted an earlier version of this macro
    (from before the column reorder), you need to replace it with the
    version below — the column positions it reads changed.
      a. Run the script once normally so a plain cuckoo_export.xlsx
         exists with the columns/formulas already in it.
      b. Open that file in Excel. Press Alt+F11 to open the VBA editor.
      c. In the Project pane on the left, double-click the entry for
         this sheet (e.g. "Sheet1 (Cuckoo Export)") — NOT "Insert >
         Module". This must be the sheet's own code-behind so the
         double-click event actually fires.
      d. Paste in (replacing any earlier version):

           Private Sub Worksheet_BeforeDoubleClick(ByVal Target As Range, Cancel As Boolean)
               Dim r As Long, num As String, msg As String, url As String
               If Target.Column <> 7 Then Exit Sub   ' column G = WhatsApp Link
               r = Target.Row
               If r < 2 Then Exit Sub
               num = Trim(Cells(r, "K").Value)       ' hidden helper column
               msg = Cells(r, "J").Value              ' WhatsApp Message
               If num = "" Or msg = "" Then Exit Sub
               ' web.whatsapp.com (not wa.me) on purpose — wa.me links often
               ' get intercepted by WhatsApp Desktop if it's installed, and
               ' the desktop app silently drops the pre-filled text for a
               ' chat that already exists. Routing through web.whatsapp.com
               ' forces it into an actual browser tab, where the text
               ' reliably shows up.
               url = "https://web.whatsapp.com/send?phone=" & num & "&text=" & WorksheetFunction.EncodeURL(msg)
               ActiveWorkbook.FollowHyperlink Address:=url, NewWindow:=True
               Cancel = True
           End Sub

      e. Close the VBA editor. File > Save As > "Excel Macro-Enabled
         Workbook (*.xlsm)" > save it as exactly
         "cuckoo_export_template.xlsm", in the same folder this script
         runs from.
      f. From then on, every run of this script detects that file and
         writes into it automatically — this setup never needs
         repeating (unless the columns change again).
    """
    ccs_rows = ccs_rows or []

    if not records:
        print("No records captured — nothing written.")
        return OUTPUT_FILE

    records = sorted(records, key=lambda r: r.get("sales_no", ""))
    filters_by_sales_no = _build_filters_lookup(ccs_rows)

    if os.path.exists(TEMPLATE_FILE):
        try:
            wb = openpyxl.load_workbook(TEMPLATE_FILE, keep_vba=True)
            ws = wb.active
            # Clear out any previous run's rows before writing fresh ones,
            # but leave row 1 (headers) and the macro itself untouched.
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            _populate_sheet(ws, records, filters_by_sales_no)
            build_route_plan_sheet(wb, records)
            wb.save(OUTPUT_FILE_XLSM)
            return OUTPUT_FILE_XLSM
        except Exception as e:
            print(f"  !! Could not write into {TEMPLATE_FILE} ({e}) — "
                  f"falling back to a plain .xlsx instead.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuckoo Export"
    _populate_sheet(ws, records, filters_by_sales_no)
    build_route_plan_sheet(wb, records)
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


if __name__ == "__main__":
    run()
