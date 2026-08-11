"""
Cuckoo+ Service Specialist — address grouping

Run this AFTER main.py (the scraper) has produced cuckoo_export.xlsm (or
.xlsx if you haven't set up the WhatsApp macro template), and BEFORE
typing in proposed_date — this script never touches proposed_date or
the WhatsApp Message/Link columns, only adds new ones, so it's safe to
run first and see the grouping before deciding on dates.

  1. Geocodes each customer's install_address using OpenStreetMap's free
     Nominatim service (no API key, no account, no cost). Results are
     cached locally in geocode_cache.json, so re-running this script
     after adding a few new customers doesn't re-geocode addresses it's
     already resolved.
  2. Groups customers by Malaysia's actual administrative hierarchy —
     state (negeri) > district (daerah) > city/town > suburb/precinct >
     neighbourhood — pulled from OpenStreetMap's own structured address
     breakdown for each geocoded point, rather than an arbitrary
     distance radius. This is what correctly separates two addresses
     that happen to be physically close but are in different named
     areas (e.g. same precinct, different named sub-development), and
     just as correctly keeps together addresses that share every level
     down to neighbourhood. No target group size or day-count — groups
     are exactly whatever the real place hierarchy says they are.
  3. Classifies each address as "Landed" or "High-Rise" from the
     address text itself (block/unit-code patterns and keywords like
     BLOK, TINGKAT, PANGSAPURI, etc. mean high-rise; anything without
     those signals is assumed landed).
  4. Writes "Area" and "Housing Type" columns onto the existing "Cuckoo
     Export" sheet — added at the END (columns K, L) specifically so
     nothing already there (the WhatsApp Message/Link formulas, the
     macro's hidden helper column) has to move or get renumbered.
  5. Adds a new "Route Plan" sheet — ONE sorted view: grouped by Area,
     and within each area split into a Landed block and a High-Rise
     block, each listing its stops with a Google Maps link per address.

Note on why the main "Cuckoo Export" sheet itself isn't physically
re-sorted by Area/Housing Type: its formulas (WhatsApp Message/Link)
are written as plain text with hardcoded row numbers, not real Excel
relative references — openpyxl has no concept of formula semantics, so
if this script moved rows around, those formulas would silently end up
pointing at the wrong row's data. The Route Plan sheet is pure values
with no formulas, so it's completely safe to sort there instead — if
you want the main sheet sorted too, Excel's own Data > Sort (done by
you, in Excel, not by this script) handles relative references
correctly and won't have this problem.

ADDRESSES THAT DON'T GEOCODE
-----------------------------
Some addresses (especially very detailed ones — specific block/unit
numbers) won't resolve on the first try. This script automatically
retries using a simplified, street/precinct-level version of the
address if the full one fails. If that still fails, the customer is
listed separately at the bottom of the Route Plan sheet under "Could
not place automatically" — you'll need to place those manually.

NOMINATIM USAGE NOTE
-----------------------------
Nominatim's usage policy asks for a maximum of 1 request/second and an
identifying User-Agent — both handled below (NOMINATIM_USER_AGENT,
REQUEST_DELAY_SECONDS). If you plan on running this regularly, consider
editing NOMINATIM_USER_AGENT to include a real contact (e.g. your
email) per their policy, though for a once-a-month run of a few dozen
addresses this is very low volume.
"""

import json
import os
import re
import time
import urllib.parse
import urllib.request

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# CONFIG
# ============================================================

# Same filenames main.py uses — this script looks for the macro-enabled
# one first (so it doesn't clobber the macro), falling back to plain xlsx.
OUTPUT_FILE_XLSM = "cuckoo_export.xlsm"
OUTPUT_FILE_XLSX = "cuckoo_export.xlsx"

GEOCODE_CACHE_FILE = "geocode_cache.json"

# A geocoding match is rejected as "too broad to be useful" if its
# bounding box spans more than this many degrees in either direction
# (roughly 0.05° ≈ 5.5km at Malaysia's latitude). This is what stops a
# detailed address that fails to parse from silently falling back to a
# whole-city or whole-state match — without this, many different
# addresses in the same city could all collapse onto one identical
# point instead of being honestly treated as failures.
MAX_MATCH_SPAN_DEGREES = 0.05

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
# Nominatim's usage policy asks for an identifying User-Agent. Feel free
# to edit this to include a real contact if you run this often.
NOMINATIM_USER_AGENT = "cuckoo-plus-route-planner/1.0 (personal workflow tool)"
# Their policy caps usage at 1 request/second — this stays comfortably
# under that.
REQUEST_DELAY_SECONDS = 1.1

AREA_COLORS = [
    "#1f4e78", "#c00000", "#2e7d32", "#e65100", "#6a1b9a",
    "#00838f", "#ad1457", "#4e342e", "#546e7a", "#9e9d24",
]


# ============================================================
# Geocoding (OpenStreetMap Nominatim, free, cached)
# ============================================================

def load_geocode_cache():
    if not os.path.exists(GEOCODE_CACHE_FILE):
        return {}
    try:
        with open(GEOCODE_CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_geocode_cache(cache):
    with open(GEOCODE_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def _nominatim_query(params):
    """One rate-limited request to Nominatim. Returns the raw first
    result dict (with lat/lon/boundingbox/address/etc.) or None."""
    params = {**params, "addressdetails": 1}
    url = NOMINATIM_URL + "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(
        url, headers={"User-Agent": NOMINATIM_USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            results = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"    [geocode] request failed: {type(e).__name__}: {e}")
        return None
    if not results:
        return None
    return results[0]


# Each entry: (canonical level name, candidate OSM/Nominatim field names
# to check, in order). Malaysia's real admin hierarchy is Negeri (state)
# > Daerah (district) > Bandar/Pekan (city/town) > Presint/Taman/Mukim
# (suburb) > a finer named pocket within that (neighbourhood) — but OSM
# contributors don't always tag things under the exact same field name,
# so each level checks a few plausible alternates.
_HIERARCHY_LEVELS = [
    ("state", ["state"]),
    ("district", ["state_district", "county"]),
    ("city", ["city", "town", "municipality"]),
    ("suburb", ["suburb", "city_district", "borough"]),
    ("neighbourhood", ["neighbourhood", "quarter", "residential", "hamlet"]),
]


def extract_hierarchy(result):
    """
    Pulls a (state, district, city, suburb, neighbourhood) tuple out of
    a Nominatim result's structured "address" breakdown — this is what
    actually drives grouping now, instead of a raw distance radius.
    Missing levels come through as None (e.g. a smaller town might have
    no separate "district" tag) — this generally still groups correctly
    with other addresses missing the same level, it just means that
    level isn't part of what distinguishes them.
    """
    address = result.get("address", {}) if result else {}
    levels = []
    for _, candidates in _HIERARCHY_LEVELS:
        value = None
        for field in candidates:
            if address.get(field):
                value = address[field]
                break
        levels.append(value)
    return tuple(levels)


def hierarchy_label(hierarchy):
    """Human-readable path, e.g. 'Selangor > Petaling > Petaling Jaya > SS2'."""
    parts = [p for p in hierarchy if p]
    return " > ".join(parts) if parts else "Unknown area"


def _is_specific_enough(result):
    """
    Rejects matches that are too broad to be useful — e.g. when a
    detailed unit/block address doesn't parse and Nominatim quietly
    falls back to matching just the city or state name. Left
    unchecked, that gives every address in the same city the SAME
    coordinates (exactly the "all the same coordinates" bug), which
    silently produces a garbage grouping instead of an honest failure.

    Uses the result's boundingbox width/height as a proxy for how
    precise the match is — a specific address's bounding box is tiny;
    a city or state's is not.
    """
    bbox = result.get("boundingbox")
    if not bbox:
        return False
    south, north, west, east = (float(x) for x in bbox)
    return (north - south) <= MAX_MATCH_SPAN_DEGREES and (east - west) <= MAX_MATCH_SPAN_DEGREES


def _simplify_to_street_level(address):
    """
    Malaysian residential addresses often lead with a unit/block/lot
    code (e.g. "A-T07-U06, BLOK A ZONE 6B ...") that essentially never
    exists in OpenStreetMap's data — there's no realistic way to
    geocode down to a specific unit, and feeding Nominatim the whole
    noisy string often makes the full-address query fail entirely.

    This strips everything before the first recognizable street/area
    keyword (JALAN, PERSIARAN, LORONG, PRESINT, TAMAN, BANDAR, LEBUH,
    LINGKARAN — all common in Malaysian addresses), keeping the part
    that's actually likely to be mapped. Returns None if none of those
    keywords appear (nothing to simplify).
    """
    keywords = r"(JALAN|PERSIARAN|LORONG|PRESINT|TAMAN|BANDAR|LEBUH|LINGKARAN)"
    match = re.search(keywords, address, re.IGNORECASE)
    if not match:
        return None
    return address[match.start():]


def geocode_address(address, cache):
    """
    Returns {"lat": ..., "lon": ..., "hierarchy": (...)} or None, using
    the cache first. On a cache miss, tries three queries in order,
    keeping the first that's actually usable:
      1. The full address, as free text.
      2. A simplified version with any unmapped unit/block prefix
         stripped off (see _simplify_to_street_level) — usually
         succeeds where #1 fails, since OSM rarely has unit-level data
         for Malaysian residential addresses, but the street/precinct
         name after it often IS mapped.
      3. A STRUCTURED query on just the postcode (a 5-digit number in
         the address) — the last resort. This is deliberately coarser
         (postcode-level, not street-level), so unlike #1 and #2 it is
         NOT rejected for being "too broad" — that check exists to
         catch a detailed address silently collapsing into a whole
         city/state match, not to second-guess a fallback that's
         supposed to be coarse.

    If none of the three give a usable result, this is treated as a
    genuine failure (None) rather than silently accepting a city/state-
    wide pin that would look "successful" but actually collapse many
    different addresses onto one identical point.
    """
    key = address.strip()
    if not key:
        return None

    if key in cache:
        cached = cache[key]
        if cached is None:
            return None  # a previous run already tried and genuinely failed
        if isinstance(cached, dict) and "hierarchy" in cached:
            return {**cached, "hierarchy": tuple(cached["hierarchy"])}
        # else: old cache format from before hierarchy tracking existed
        # (just [lat, lon], no address breakdown) — fall through and
        # re-geocode, since grouping now needs the hierarchy data.

    time.sleep(REQUEST_DELAY_SECONDS)
    result = _nominatim_query({
        "q": key,
        "format": "jsonv2",
        "countrycodes": "my",
        "limit": 1,
    })
    if result is not None and not _is_specific_enough(result):
        result = None

    if result is None:
        simplified = _simplify_to_street_level(key)
        if simplified:
            time.sleep(REQUEST_DELAY_SECONDS)
            result = _nominatim_query({
                "q": simplified,
                "format": "jsonv2",
                "countrycodes": "my",
                "limit": 1,
            })
            if result is not None and not _is_specific_enough(result):
                result = None

    if result is None:
        postcode_match = re.search(r"\b(\d{5})\b", key)
        if postcode_match:
            time.sleep(REQUEST_DELAY_SECONDS)
            # No _is_specific_enough check here on purpose — see docstring.
            result = _nominatim_query({
                "postalcode": postcode_match.group(1),
                "country": "Malaysia",
                "format": "jsonv2",
                "limit": 1,
            })

    if result:
        hierarchy = extract_hierarchy(result)
        geocoded = {
            "lat": float(result["lat"]),
            "lon": float(result["lon"]),
            "hierarchy": hierarchy,
        }
        print(
            f"    [geocode] \"{key[:40]}...\" -> {hierarchy_label(hierarchy)}")
        cache[key] = {"lat": geocoded["lat"],
                      "lon": geocoded["lon"], "hierarchy": list(hierarchy)}
    else:
        geocoded = None
        cache[key] = None

    save_geocode_cache(cache)
    return geocoded


# ============================================================
# Google Maps link (plain URL — no API key needed)
# ============================================================

def gmaps_single_link(lat, lon):
    return f"https://www.google.com/maps/search/?api=1&query={lat},{lon}"


# ============================================================
# Landed vs high-rise classification (from the address text itself,
# no geocoding needed)
# ============================================================

_HIGH_RISE_KEYWORDS = re.compile(
    r"\b(BLOK|BLOCK|TINGKAT|PANGSAPURI|KONDOMINIUM|CONDOMINIUM|CONDO|"
    r"APARTMENT|FLAT|RESIDENSI|RESIDENCE|SUITES?|MENARA|TOWER|PARCEL)\b",
    re.IGNORECASE,
)
# Catches block-floor-unit style codes like "A-12-05" or "A-T13-U07"
# even when none of the keywords above happen to be present.
_UNIT_CODE_PATTERN = re.compile(
    r"\b[A-Z]-[A-Z]?\d+[A-Z]?-[A-Z]?\d+\b", re.IGNORECASE)


def classify_housing_type(address):
    """
    "Landed" or "High-Rise", guessed from the address text alone —
    no geocoding involved, so this works even for addresses that
    couldn't be placed on a map. Malaysian apartment/condo addresses
    reliably include either an explicit building-type keyword (BLOK,
    TINGKAT, PANGSAPURI, ...) or a block-floor-unit code (A-12-05,
    A-T13-U07). Anything with neither signal is assumed landed — a
    standalone house address is just a house number and street, with
    nothing more specific to detect.
    """
    if _HIGH_RISE_KEYWORDS.search(address) or _UNIT_CODE_PATTERN.search(address):
        return "High-Rise"
    return "Landed"


# ============================================================
# Street-name extraction (from the address text, no geocoding needed)
# ============================================================

_STREET_PATTERN = re.compile(
    r"\b(?:JALAN|JLN|LORONG|PERSIARAN|LEBUH|LINGKARAN)\s+"
    r"[A-Z0-9/.\-]+(?:\s+[A-Z0-9/.\-]+)?",
    re.IGNORECASE,
)


def extract_street(address):
    """
    Pulls out the road name/code (e.g. "JALAN P11A 1/1", "JLN TANGGILAN
    11D/17") directly from the address text. This exists because a lot
    of Putrajaya's internal residential lanes simply aren't in
    OpenStreetMap's data at all — Nominatim can only place them at
    precinct level, which is genuinely the finest grouping it can
    support for those cases (see geocode_address's docstring). But the
    actual street name is still sitting right there in the text even
    when it can't be geocoded, so it's used to sort stops WITHIN an
    area/precinct — same-lane houses end up next to each other in the
    Route Plan sheet instead of in random order, without needing to
    split the precinct into a separate top-level Area for every street
    (which would fragment things into unusably tiny groups).
    Returns "" if no street keyword is found.
    """
    match = _STREET_PATTERN.search(address)
    if not match:
        return ""
    return match.group(0).strip().upper()


def _street_sort_key(street):
    """Normalizes minor spacing inconsistencies (e.g. "P11J" vs "P11 J",
    which really are the same road) so they sort/group together instead
    of being treated as different streets purely due to a stray space."""
    return street.replace(" ", "")


# ============================================================
# Main
# ============================================================

def find_input_file():
    if os.path.exists(OUTPUT_FILE_XLSM):
        return OUTPUT_FILE_XLSM
    if os.path.exists(OUTPUT_FILE_XLSX):
        return OUTPUT_FILE_XLSX
    return None


def run():
    target_file = find_input_file()
    if not target_file:
        print(f"Couldn't find {OUTPUT_FILE_XLSM} or {OUTPUT_FILE_XLSX} — "
              f"run main.py first to generate the export.")
        return

    is_xlsm = target_file.endswith(".xlsm")
    wb = openpyxl.load_workbook(target_file, keep_vba=is_xlsm)
    ws = wb.active  # "Cuckoo Export" sheet

    # Column A = sales_no, D = install_contact_person, E = install_mobile1,
    # F = install_address — matching main.py's DATA_COLUMNS order.
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        sales_no = ws.cell(row=row_idx, column=1).value
        if not sales_no:
            continue
        rows.append({
            "row_idx": row_idx,
            "sales_no": sales_no,
            "contact": ws.cell(row=row_idx, column=4).value or "",
            "phone": ws.cell(row=row_idx, column=5).value or "",
            "address": ws.cell(row=row_idx, column=6).value or "",
        })

    if not rows:
        print("No customer rows found in the sheet — nothing to plan.")
        return

    for r in rows:
        r["housing_type"] = classify_housing_type(r["address"])
        r["street"] = extract_street(r["address"])

    print(f"Geocoding {len(rows)} address(es) (cached ones are instant, "
          f"new ones take ~1s each)...")
    cache = load_geocode_cache()
    geocoded, failed = [], []
    for i, r in enumerate(rows, start=1):
        result = geocode_address(r["address"], cache)
        if result:
            r["lat"], r["lon"], r["hierarchy"] = result["lat"], result["lon"], result["hierarchy"]
            geocoded.append(r)
        else:
            failed.append(r)
        print(f"  [{i}/{len(rows)}] {r['sales_no']}: "
              f"{'OK' if result else 'could not place'}")

    if not geocoded:
        print("Nothing could be geocoded — check your internet connection "
              "and try again.")
        return

    areas = {}
    for r in geocoded:
        areas.setdefault(r["hierarchy"], []).append(r)

    # Sort by the hierarchy tuple itself — this naturally nests: all of
    # one state sorts together, then within it all of one district, and
    # so on down to neighbourhood. None (a level that wasn't available
    # for a given match) sorts as "" so it doesn't crash comparing
    # against real strings, and groups those cases first within a level.
    def sort_key(hierarchy):
        return tuple(level or "" for level in hierarchy)
    sorted_hierarchies = sorted(areas.keys(), key=sort_key)

    # --- Write Area + Housing Type onto the existing sheet, columns K/L ---
    FONT = "Arial"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, header in [(11, "Area"), (12, "Housing Type")]:
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = header_fill

    for r in geocoded:
        ws.cell(row=r["row_idx"], column=11,
                value=hierarchy_label(r["hierarchy"])).font = Font(name=FONT, size=10)
        ws.cell(row=r["row_idx"], column=12,
                value=r["housing_type"]).font = Font(name=FONT, size=10)
    for r in failed:
        ws.cell(row=r["row_idx"], column=11,
                value="?").font = Font(name=FONT, size=10)
        ws.cell(row=r["row_idx"], column=12,
                value=r["housing_type"]).font = Font(name=FONT, size=10)

    ws.column_dimensions["K"].width = 30
    ws.column_dimensions["L"].width = 14

    # --- Route Plan sheet: ONE sorted view — grouped by Area, and within
    # each area split into a Landed block then a High-Rise block. This
    # sheet is pure values (no formulas), so unlike the main sheet it's
    # completely safe to lay out in whatever order is most useful. ---
    if "Route Plan" in wb.sheetnames:
        del wb["Route Plan"]
    rp = wb.create_sheet("Route Plan")

    subheader_fill = PatternFill("solid", fgColor="D9E1F2")
    r_idx = 1

    def write_stop_block(rows_for_block, start_row):
        row = start_row
        headers = ["Sales No", "Contact", "Phone",
                   "Street", "Address (tap to open)"]
        for col_idx, h in enumerate(headers, start=1):
            c = rp.cell(row=row, column=col_idx, value=h)
            c.font = Font(name=FONT, size=10, bold=True)
            c.fill = subheader_fill
        row += 1
        # Sorting by (normalized street, sales_no) clusters same-lane
        # houses together even within a precinct that couldn't be
        # geocoded any finer than "Presint 11" as a whole — see
        # extract_street()'s docstring for why this exists.
        sorted_rows = sorted(
            rows_for_block,
            key=lambda r: (_street_sort_key(r["street"]), r["sales_no"]),
        )
        for r in sorted_rows:
            rp.cell(row=row, column=1, value=r["sales_no"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=row, column=2, value=r["contact"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=row, column=3, value=r["phone"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=row, column=4, value=r["street"] or "?").font = Font(
                name=FONT, size=10)
            addr_cell = rp.cell(row=row, column=5, value=r["address"])
            addr_cell.hyperlink = gmaps_single_link(r["lat"], r["lon"])
            addr_cell.font = Font(name=FONT, size=10,
                                  color="1155CC", underline="single")
            addr_cell.alignment = Alignment(wrap_text=True)
            row += 1
        return row

    for area_idx, hierarchy in enumerate(sorted_hierarchies):
        area_rows = areas[hierarchy]
        area_color = AREA_COLORS[area_idx % len(AREA_COLORS)]
        cell = rp.cell(row=r_idx, column=1,
                       value=f"{hierarchy_label(hierarchy)} — {len(area_rows)} address(es)")
        cell.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=area_color.lstrip("#"))
        rp.merge_cells(start_row=r_idx, start_column=1,
                       end_row=r_idx, end_column=5)
        r_idx += 1

        landed = [r for r in area_rows if r["housing_type"] == "Landed"]
        high_rise = [r for r in area_rows if r["housing_type"] == "High-Rise"]

        for label, subset in [("Landed", landed), ("High-Rise", high_rise)]:
            if not subset:
                continue
            sub_cell = rp.cell(row=r_idx, column=1,
                               value=f"{label} ({len(subset)})")
            sub_cell.font = Font(name=FONT, size=11, bold=True, italic=True)
            rp.merge_cells(start_row=r_idx, start_column=1,
                           end_row=r_idx, end_column=5)
            r_idx += 1
            r_idx = write_stop_block(subset, r_idx)
            r_idx += 1  # blank row after each sub-block

        r_idx += 1  # blank row between areas

    if failed:
        cell = rp.cell(row=r_idx, column=1,
                       value="Could not place automatically — assign manually")
        cell.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        rp.merge_cells(start_row=r_idx, start_column=1,
                       end_row=r_idx, end_column=5)
        r_idx += 1
        headers = ["Sales No", "Contact", "Phone",
                   "Street", "Address", "Housing Type"]
        for col_idx, h in enumerate(headers, start=1):
            c = rp.cell(row=r_idx, column=col_idx, value=h)
            c.font = Font(name=FONT, size=10, bold=True)
            c.fill = subheader_fill
        r_idx += 1
        for r in sorted(failed, key=lambda r: (_street_sort_key(r["street"]), r["sales_no"])):
            rp.cell(row=r_idx, column=1, value=r["sales_no"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=r_idx, column=2, value=r["contact"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=r_idx, column=3, value=r["phone"]).font = Font(
                name=FONT, size=10)
            rp.cell(row=r_idx, column=4, value=r["street"] or "?").font = Font(
                name=FONT, size=10)
            addr_cell = rp.cell(row=r_idx, column=5, value=r["address"])
            addr_cell.alignment = Alignment(wrap_text=True)
            addr_cell.font = Font(name=FONT, size=10)
            rp.cell(row=r_idx, column=6, value=r["housing_type"]).font = Font(
                name=FONT, size=10)
            r_idx += 1

    rp.column_dimensions["A"].width = 14
    rp.column_dimensions["B"].width = 24
    rp.column_dimensions["C"].width = 16
    rp.column_dimensions["D"].width = 22
    rp.column_dimensions["E"].width = 50
    rp.column_dimensions["F"].width = 14

    wb.save(target_file)

    print(f"\nDone. {len(geocoded)} address(es) grouped into {len(areas)} area(s), "
          f"{len(failed)} address(es) need manual placement.")
    print(f"Written back to {target_file} — see the 'Route Plan' sheet (sorted by "
          f"Area, then Landed/High-Rise) and the new Area/Housing Type columns "
          f"on 'Cuckoo Export'.")


if __name__ == "__main__":
    run()
